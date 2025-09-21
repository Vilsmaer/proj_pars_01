# format_to_excel.py — Преобразуем CSV в красивый Excel

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# === КОНФИГУРАЦИЯ ===
INPUT_CSV = r"D:\ANKO\proj_pars_01\all_products.csv"
OUTPUT_XLSX = r"D:\ANKO\proj_pars_01\products.xlsx"

# Цвета (опционально)
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # темно-синий
HEADER_FONT = Font(bold=True, color="FFFFFF")  # белый жирный
EVEN_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # светло-серый

# === ЧИТАЕМ CSV ===
print(" Читаем CSV...")
df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")

# Удаляем дубликаты по URL (если есть)
df = df.drop_duplicates(subset=['url'], keep='first').reset_index(drop=True)

print(f" Загружено {len(df)} уникальных товаров")

# Сохраняем во временный Excel
df.to_excel(OUTPUT_XLSX, index=False, sheet_name="Товары")

# === ФОРМАТИРУЕМ EXCEL ===
print(" Форматируем Excel...")

wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

# 1. Жирные заголовки + цвет фона
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 2. Автоподбор ширины столбцов
for col in range(1, ws.max_column + 1):
    column_letter = get_column_letter(col)
    max_length = 0
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    adjusted_width = min(max_length + 2, 50)  # ограничим максимумом 50
    ws.column_dimensions[column_letter].width = adjusted_width

# 3. Добавляем фильтры
ws.auto_filter.ref = ws.dimensions

# 4. (Опционально) Заливка чётных строк
for row in range(2, ws.max_row + 1):
    if row % 2 == 0:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = EVEN_ROW_FILL

# Сохраняем
wb.save(OUTPUT_XLSX)
print(f" Файл сохранён: {OUTPUT_XLSX}")
print(" Готово! Открой файл в Excel — всё красиво и удобно!")